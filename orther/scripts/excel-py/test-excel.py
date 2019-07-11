#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author  : Swimming

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


# A to 1 ,2 to B
from openpyxl.utils import get_column_letter, column_index_from_string

tmp_file = '/Users/swimming/Desktop/工作/05-06_06_day.xlsx'

wb = load_workbook(tmp_file)
ws = wb.get_active_sheet

for i in range(1, ws.max_column+1):
    ft = ws.cell(row=1, column=i)
    ft.font = Font(bold=True, size="14", color="FFFFFFFF")
    ft.alignment = Alignment(horizontal='center', vertical='center')
    ft.fill = PatternFill('solid', fgColor='104E8B')
    # 1874CD

    ft_all = ws.column_dimensions[get_column_letter(i)]
    if i == 1:
        ft_all.width = 25
    else:
        ft_all.width = 12
        # ft_all.alignment = Alignment(vertical='center', horizontal='center')
        # ws[get_column_letter(i)].alignment = Alignment(vertical='center', horizontal='center')
        for x in range(1, ws.max_row + 1):
            ws[get_column_letter(i) + str(x)].alignment = Alignment(horizontal='center')

for row_i in range(1,ws.max_row + 1):
    ws.row_dimensions[row_i].height = 20

# 冻结
ws.freeze_panes = ws['A2']
wb.save(filename = tmp_file)
wb.close()

