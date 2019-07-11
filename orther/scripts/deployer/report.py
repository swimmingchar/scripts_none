#!/usr/bin/env python
# coding:utf-8
import time
import jenkins
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# 分组
all_groups = dict()
## 前端组
h5_Group = "app-ykb,app-agent,app-merchant,	app-agent-android,app-merchant-android".split(',')
## 架构组
arch_Group = "bwf-console,bwf-core,bwf-in,cas,job-dispatcher,position-core,sms-core,tp-auth,eureka,jwt-ca,monitor-turbine,stp-portal".split(',')
## 交易组
trade_Group = "trade-core-card,trade-core-service,trade-data-sync,trade-prepose-unionpay,trade-security,trade-tms,trade-tms-console".split(',')
## 账户组
account_Group = "account-core,account-portal,recon-core,recon-portal,remit-core,remit-portal,settle-core,settle-portal," \
                "remit-front,remit-gateway".split(',')
## 风控组
risk_Group = "risk-manage-core,risk-manage-task,risk-sync,rule-engine-core,rule-engine-portal,trans-route-core," \
             "trans-route-portal,trans-route-sync,trans-route-task,markting-client,markting-sync," \
             "profit-core,recommed-core,profit-portal,profit-task,markting-portal,recommed-portal,recommed-task," \
             "markting-task".split(',')
## 运营保障组
oper_Group = "agentserver-core,agentserver-gateway,agentserver-portal,agentserver-query,agentserver-web," \
             "custserver-core,custserver-gateway,custserver-portal,custserver-query,manage-core,manage-sync," \
             "manage-task,process-server,terminal-server-core,terminal-server-portal,terminal-server-query,boss-core," \
             "boss-out,boss-portal,sales-portal".split(',')

all_groups["架构组"] = arch_Group
all_groups["交易组"] = trade_Group
all_groups["清算组"] = account_Group
all_groups["风控组"] = risk_Group
all_groups["前端组"] = h5_Group
all_groups["运营保障组"] = oper_Group

single_app = list("custserver-sync,markting-sync,risk-sync,terminal-server-sync,trade-data-sync,trans-route-sync,manage-sync,recon-core".split(","))

def get_group(all_g, app):
    return ''.join([k for k, v in all_g.items() if app in v])


def stats_jenkins():
    server = jenkins.Jenkins('http://10.10.10.25:8080/', username='admin', password='zhangjian')
    return server


def format_excel(excel_ws,int):

    for excel_i in range(1, excel_ws.max_column + 1):
        ft = excel_ws.cell(row=1, column = excel_i)
        ft.font = Font(bold=True, size="14", color="FFFFFFFF")
        ft.alignment = Alignment(horizontal='center', vertical='center')
        ft.fill = PatternFill('solid', fgColor="1874CD")
        # 1874CD,104E8B

        ft_all = excel_ws.column_dimensions[get_column_letter(excel_i)]
        if excel_i == 1:
            ft_all.width = 24
        if excel_i != 1:
            ft_all.width = 16
            for excel_x in range(1, excel_ws.max_row + 1):
                excel_ws[get_column_letter(excel_i) + str(excel_x)].alignment = Alignment(horizontal='center')

    for row_i in range(1, excel_ws.max_row + 1):
        excel_ws.row_dimensions[row_i].height = 20
    # 冻结
    excel_ws.freeze_panes = excel_ws['A2']

in_year = time.strftime("%Y")

m_d = input("请输入月日按回车键,eg:04-23:")
on_day = input("请输入之后的天数，默认7天：")
day = 0
if len(on_day) == 0:
    day = 7*86400
    on_day = 7
else:
    day = int(on_day) * 86400

in_month = m_d[0:2]
in_day = m_d[-2:]
start_time = in_year + '-' + in_month + '-' + in_day + ' 00:00:00'
start_time = int(time.mktime(time.strptime(start_time,"%Y-%m-%d %H:%M:%S")))
end_time = start_time + int(day)
# 获取job列表
j_server = stats_jenkins()
# 部署列表
job_list = list()
# 重启列表
reboot_list = list()

for item in j_server.get_all_jobs():
    app_name = item['name']
    if app_name.endswith("_Reboot"):
        reboot_list.append(app_name)
    else:
        job_list.append(app_name)

job_xlwt = list()
rejob_xlwt = list()

# 获取应用在时间范围内的构件清单
for job in job_list:
    if job.endswith("RollBack"):
        continue
    try:
        gj_number = j_server.get_job_info(job)['builds'][0]['number']
    except Exception as msg:
        print("构件版本号获取错误,%s"% msg)
        continue
    # print(job)
    # 循环app内构件数
    for num in range(gj_number, 0, -1):
        build_info = j_server.get_build_info(name=job, number=num)
        gj_time = int(str(build_info['timestamp'])[:10])
        if gj_time >= start_time:
            if gj_time <= end_time:
                user = build_info["actions"][1]['causes'][0]['userName']
                res = build_info['result']
                svn_num = None
                if res.startswith("S"):
                    for x in build_info['changeSet']['revisions']:
                        if svn_num != None :
                            if job == x['module'].split('/')[-1].strip() or x['module'].split('/')[-1].strip() == 'trunk':
                                svn_num = x['module'].split('/')[-1].strip() + '-' + str(x['revision']).strip() + '||' + svn_num
                            else:
                                svn_num = svn_num + '||' + x['module'].split('/')[-1].strip() + '-' + str(x['revision']).strip()
                        if svn_num == None:
                            svn_num = (x['module'].split('/')[-1]).strip() + '-' + (str(x['revision'])).strip()
                # H5 版本号，适用于单一版本号
                if job in h5_Group:
                    svn_num = build_info['changeSet']['revisions'][0]['revision']
                # 部署工作IP获取
                job_ip = build_info['actions'][0]['parameters'][1]['value']
                job_xlwt.append((job, time.strftime("%m/%d %H:%M:%S", time.localtime(gj_time)), user, job_ip ,res, svn_num, num,
                                 get_group(all_groups, job)))
            else:
                continue
        else:
            break

# 重启应用明细获取
for r_job in reboot_list:
    try:
        rgj_num = j_server.get_job_info(r_job)['builds'][0]['number']
    except Exception as msg:
        # print("重启JOB获取构件号错误，%s" % msg)
        continue
    # print(r_job)
    # 获取详细内容
    for rnum in range(rgj_num, 0, -1):
        build_info = j_server.get_build_info(name=r_job, number=rnum)
        rgj_time = int(str(build_info['timestamp'])[:10])
        if rgj_time >= start_time:
            if rgj_time <= end_time:
                user = build_info['actions'][1]['causes'][0]['userName']
                res = build_info['result']
                job = r_job.split("_Reboot")[0]
                job_rip = build_info['actions'][0]['parameters'][1]['value']
                rejob_xlwt.append((job, time.strftime("%m/%d %H:%M:%S", time.localtime(rgj_time)), user, job_rip ,res, rnum,
                                   get_group(all_groups, job)))
            else:
                continue
        else:
            break

# 表格内容填写
wb = Workbook()
# 重启明细
reboot_ws = wb.create_sheet("应用重启明细")
reboot_ws.append(('应用名', '重启时间', '重启用户','IP地址','重启结果结果', 'JOB构件号', '属组'))
for i in rejob_xlwt:
    reboot_ws.append(i)

format_excel(reboot_ws,1)

# 重启汇总
reboot_all_ws = wb.create_sheet("应用重启汇总", 0)
reboot_all_ws.append(("应用名", "重启成功次数", "重启失败次数", "取消重启次数", "属组"))
reboot_job = dict()
for r_j in rejob_xlwt:
    job_name = r_j[0]
    # job_group = r_j[-1]
    if job_name not in reboot_job.keys():
        reboot_job[job_name] = {'success': 0, 'fail': 0, 'cancel': 0}
    # print("Reboot JOB NAME LIST %s"%r_j)
    if r_j[4].startswith("S"):
        reboot_job[job_name]['success'] = reboot_job[job_name]['success'] + 1
    if r_j[4].startswith("F"):
        reboot_job[job_name]['fail'] = reboot_job[job_name]['fail'] + 1
    if r_j[4].startswith("AB"):
        reboot_job[job_name]['cancel'] = reboot_job[job_name]['cancel'] + 1
    reboot_job[job_name]['group'] = r_j[-1]

for jitme in reboot_job:
    reboot_all_ws.append((jitme, reboot_job[jitme]['success'], reboot_job[jitme]['fail'], reboot_job[jitme]['cancel'],
                          reboot_job[jitme]['group']))

format_excel(reboot_all_ws,0)
# ---------------------+++++++++-----------------------
# 部署
ws = wb.create_sheet("应用部署明细", 1)
# 删除默认多余sheet工作簿
try:
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
except Exception as msg:
    print("删除原始表格失败，请忽略！")

# 添加部署明细
ws.append(('应用名', '构件时间', '操作人', 'IP地址' ,'构件结果', '构件版本号', 'JOB构件号', '属组'))
for litme in job_xlwt:
    ws.append(litme)
format_excel(ws,1)

# 汇总
ws1 = wb.create_sheet('部署汇总', 0)
ws1.append(("应用名", "成功次数", "失败次数", "取消次数", "属组"))
job1 = dict()
for job in job_xlwt:
    job_name = job[0]
    job_numb = 1
    if job_name not in job1.keys():
        job1[job_name] = {'success': 0, 'fail': 0, 'cancel': 0}
    # print("JOB NAME LIST %s"%job1)
    if job[4].startswith("S"):
        if job_name in single_app:
            job_numb = 2
        job1[job_name]['success'] = job1[job_name]['success'] + job_numb
    if job[4].startswith("F"):
        job1[job_name]['fail'] = job1[job_name]['fail'] + 1
    if job[4].startswith("AB"):
        job1[job_name]['cancel'] = job1[job_name]['cancel'] + 1
    job1[job_name]['group'] = job[-1]

for jitme in job1:
    ws1.append((jitme, job1[jitme]['success']/2, job1[jitme]['fail'], job1[jitme]['cancel'], job1[jitme]['group']))

format_excel(ws1,0)

# print(job1)
ws_all = wb.create_sheet('汇总表', 0)
ws_all.append(("属组", "部署成功次数", "部署失败次数", "部署取消次数"))
groups_sum = dict()
for item in job1:
    group_name = job1[item]['group']
    if group_name not in groups_sum.keys():
        groups_sum[group_name] = {'suc': 0, 'fail': 0, 'cancel': 0}
    groups_sum[group_name]['suc'] = groups_sum[group_name]['suc'] + job1[item]['success']
    groups_sum[group_name]['fail'] = groups_sum[group_name]['fail'] + job1[item]['fail']
    groups_sum[group_name]['cancel'] = groups_sum[group_name]['cancel'] + job1[item]['cancel']
for gitem in groups_sum:
    ws_all.append((gitem, groups_sum[gitem]['suc'], groups_sum[gitem]['fail'], groups_sum[gitem]['cancel']))

ws_all.append(("汇总", "=SUM(B2:B6)", "=SUM(C2:C6)", "=SUM(D2:D6)"))


wb.save(filename=m_d + '_' + str(on_day) + '_day.xlsx')
wb.close()